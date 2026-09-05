import html
import io
import json
import os
import re
import zipfile

import pandas as pd
import streamlit as st
from docx import Document
from docxcompose.composer import Composer
from docxtpl import DocxTemplate
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from pdf_export import docx_bytes_to_pdf, merge_pdfs


DEFAULT_TEMPLATE = "内审员证书.docx"
STANDARDS_FILE = "standards.json"

FALLBACK_STANDARD_PRESETS = {
    "质量管理体系（ISO 9001:2015）": "ISO 9001:2015",
    "质量管理体系新版（ISO 9001:2026）": "ISO 9001:2026",
    "环境管理体系（ISO 14001:2015）": "ISO 14001:2015",
    "环境管理体系（ISO 14001:2026）": "ISO 14001:2026",
    "职业健康安全管理体系（ISO 45001:2018）": "ISO 45001:2018",
    "食品安全管理体系（ISO 22000:2018）": "ISO 22000:2018",
    "食品安全体系认证（FSSC 22000 V6.0）": "FSSC 22000 V6.0",
    "食品安全体系认证（FSSC 22000 V7）": "FSSC 22000 V7",
    "危害分析与关键控制点（HACCP V1.0）": "HACCP V1.0",
    "信息技术服务管理体系（ISO/IEC 20000-1:2018）": "ISO/IEC 20000-1:2018",
    "信息安全管理体系（ISO/IEC 27001:2022）": "ISO/IEC 27001:2022",
    "合规管理体系（GB/T 35770—2022/ISO 37301:2021）": "GB/T 35770—2022/ISO 37301:2021",
    "能源管理体系（ISO 50001:2018）": "ISO 50001:2018",
    "汽车行业质量管理体系（IATF 16949:2016）": "IATF 16949:2016",
    "医疗器械质量管理体系（ISO 13485:2016）": "ISO 13485:2016",
    "设施管理体系（ISO 41001:2018）": "ISO 41001:2018",
    "人工智能管理体系（ISO/IEC 42001:2023）": "ISO/IEC 42001:2023",
    "企业诚信管理体系（GB/T 31950-2023）": "GB/T 31950-2023",
    "碳管理体系（T/CCAA 39—2022）": "T/CCAA 39—2022",
    "商品售后服务评价体系（GB/T 27922-2011）": "GB/T 27922-2011",
    "测量管理体系（GB/T 19022-2003/ISO 10012:2003）": "GB/T 19022-2003/ISO 10012:2003",
    "测量管理体系新版（ISO 10012:2026）": "ISO 10012:2026",
    "物业服务（GB/T 20647.9-2006）": "GB/T 20647.9-2006",
    "BRCGS 食品安全（Food Safety Issue 9）": "BRCGS Food Safety Issue 9",
    "社会责任（SA8000:2026）": "SA8000:2026",
    "创新管理体系（ISO 56001:2024）": "ISO 56001:2024",
    "供应链安全管理体系（ISO 28000:2022）": "ISO 28000:2022",
    "水效率管理体系（ISO 46001:2019）": "ISO 46001:2019",
    "全球良好农业规范（GLOBALG.A.P. IFA v6）": "GLOBALG.A.P. IFA v6",
    "中国良好农业规范（ChinaGAP / GB/T 20014系列）": "ChinaGAP / GB/T 20014系列",
    "反贿赂管理体系（ISO 37001:2025）": "ISO 37001:2025",
    "社会责任管理体系（GB/T 39604-2020）": "GB/T 39604-2020",
    "有机产品（GB/T 19630-2019）": "GB/T 19630-2019",
    "企业知识产权合规管理体系（GB/T 29490-2023）": "GB/T 29490-2023",
    "业务连续性管理体系（GB/T 30146-2023/ISO 22301:2019）": "GB/T 30146-2023/ISO 22301:2019",
}


def load_standard_presets():
    """Load editable standard presets, falling back to built-in values."""
    try:
        with open(STANDARDS_FILE, "r", encoding="utf-8") as standards_file:
            presets = json.load(standards_file)
        if isinstance(presets, dict) and presets:
            return {str(label): str(value) for label, value in presets.items()}
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        pass
    return FALLBACK_STANDARD_PRESETS


def clean_value(value):
    """Convert spreadsheet values to clean strings without leaking NaN."""
    if value is None or pd.isna(value):
        return ""
    cleaned = str(value).strip()
    return "" if cleaned.lower() == "nan" else cleaned


def mask_id_card(id_str):
    """Mask the middle eight digits of an 18-character mainland ID number."""
    id_str = clean_value(id_str)
    if len(id_str) == 18:
        return f"{id_str[:6]}********{id_str[14:]}"
    return id_str


def parse_people(raw_people):
    """Parse names or Excel-style name/ID rows and de-duplicate by name."""
    people = []
    seen_names = set()

    for raw_line in re.split(r"[\r\n]+", raw_people or ""):
        line = raw_line.strip()
        if not line:
            continue

        tab_values = [value.strip() for value in line.split("\t")]
        if len(tab_values) >= 2:
            candidates = [(tab_values[0], tab_values[1])]
        else:
            candidates = [
                (value.strip(), "")
                for value in re.split(r"[,，;；]+", line)
                if value.strip()
            ]

        for name, id_card in candidates:
            if name in {"姓名", "名字"} and (
                not id_card or id_card in {"身份证号", "身份证号码"}
            ):
                continue
            if name and name not in seen_names:
                people.append({"姓名": name, "身份证号": id_card})
                seen_names.add(name)

    return people


def format_training_date(start_date, end_date=None):
    """Format one date or an inclusive date range for the certificate."""
    if not start_date:
        return ""
    if not end_date or end_date == start_date:
        return f"{start_date.year}年{start_date.month}月{start_date.day}日"
    if end_date < start_date:
        start_date, end_date = end_date, start_date
    if start_date.year == end_date.year and start_date.month == end_date.month:
        return (
            f"{start_date.year}年{start_date.month}月"
            f"{start_date.day}—{end_date.day}日"
        )
    if start_date.year == end_date.year:
        return (
            f"{start_date.year}年{start_date.month}月{start_date.day}日—"
            f"{end_date.month}月{end_date.day}日"
        )
    return (
        f"{start_date.year}年{start_date.month}月{start_date.day}日—"
        f"{end_date.year}年{end_date.month}月{end_date.day}日"
    )


def merge_unique(values):
    """Remove blanks and duplicates while preserving selection order."""
    result = []
    seen = set()
    for value in values:
        cleaned = clean_value(value)
        if cleaned and cleaned not in seen:
            result.append(cleaned)
            seen.add(cleaned)
    return result


def clear_quick_standards(standard_labels):
    """Clear standard checkboxes and the custom standard input."""
    for label in standard_labels:
        st.session_state[f"standard_option_{label}"] = False
    st.session_state["quick_custom_standard"] = ""


def clear_people_input():
    """Clear the quick-entry people input before widgets are rendered."""
    st.session_state["quick_people_input"] = ""


def standard_chips_html(values):
    """Build selected standards as compact, escaped visual tags."""
    chips = "".join(
        f'<span class="standard-chip">{html.escape(value)}</span>' for value in values
    )
    return f'<div class="chip-row">{chips}</div>'


def section_title(text):
    """Render a card heading without Markdown anchor icons."""
    st.markdown(
        f'<div class="panel-title">{html.escape(text)}</div>', unsafe_allow_html=True
    )


def render_checklist(items):
    """Render the generation prerequisites as a compact status card."""
    rows = []
    for complete, text in items:
        status_class = "check-ok" if complete else "check-pending"
        icon = "✓" if complete else "○"
        rows.append(
            f'<div class="check-row {status_class}"><span>{icon}</span>'
            f'<span>{html.escape(text)}</span></div>'
        )
    st.markdown(
        f'<div class="check-card">{"".join(rows)}</div>', unsafe_allow_html=True
    )


def safe_filename(value):
    """Make a short Windows-compatible filename component."""
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", clean_value(value))
    cleaned = cleaned.rstrip(". ")
    return cleaned[:80] or "未命名"


def read_template_bytes(template_source):
    if isinstance(template_source, (str, os.PathLike)):
        with open(template_source, "rb") as template_file:
            return template_file.read()
    if hasattr(template_source, "getvalue"):
        return template_source.getvalue()
    template_source.seek(0)
    return template_source.read()


def generate_documents(template_bytes, records, progress_callback=None):
    """Render one certificate per valid row and return merged DOCX, PDF, ZIP."""
    master_doc = None
    composer = None
    valid_count = 0
    pdf_pages = []
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for index, row in enumerate(records):
            name = clean_value(row.get("姓名"))
            if not name:
                continue

            doc = DocxTemplate(io.BytesIO(template_bytes))
            doc.render(
                {
                    "number": clean_value(row.get("证书编号")),
                    "name": name,
                    "id_card": mask_id_card(row.get("身份证号")),
                    "date": clean_value(row.get("培训日期")),
                    "standards": clean_value(row.get("标准号")),
                }
            )

            certificate_buffer = io.BytesIO()
            doc.save(certificate_buffer)
            certificate_bytes = certificate_buffer.getvalue()
            valid_count += 1
            stem = f"{valid_count:03d}_{safe_filename(name)}"

            archive.writestr(f"{stem}.docx", certificate_bytes)
            try:
                pdf_bytes = docx_bytes_to_pdf(certificate_bytes)
                pdf_pages.append(pdf_bytes)
                archive.writestr(f"{stem}.pdf", pdf_bytes)
            except Exception:
                pass

            current_doc = Document(io.BytesIO(certificate_bytes))
            if master_doc is None:
                master_doc = current_doc
                composer = Composer(master_doc)
            else:
                master_doc.add_page_break()
                composer.append(current_doc)

            if progress_callback:
                progress_callback((index + 1) / len(records))

    if not composer or valid_count == 0:
        return None

    merged_buffer = io.BytesIO()
    composer.save(merged_buffer)
    return {
        "count": valid_count,
        "merged": merged_buffer.getvalue(),
        "pdf": merge_pdfs(pdf_pages) if pdf_pages else None,
        "zip": zip_buffer.getvalue(),
    }


def make_excel_template():
    example_data = {
        "证书编号": ["T-2026-001（示例）"],
        "姓名": ["张三（示例）"],
        "身份证号": ["440683199001010001"],
        "培训日期": ["2026年9月1—3日"],
        "标准号": ["ISO 9001:2015、ISO 14001:2015"],
    }
    example_df = pd.DataFrame(example_data)
    template_buffer = io.BytesIO()

    with pd.ExcelWriter(template_buffer, engine="openpyxl") as writer:
        example_df.to_excel(writer, index=False, sheet_name="Sheet1")
        worksheet = writer.sheets["Sheet1"]
        for column_index, column in enumerate(example_df.columns, start=1):
            column_letter = get_column_letter(column_index)
            max_length = max(example_df[column].astype(str).map(len).max(), len(column)) + 5
            worksheet.column_dimensions[column_letter].width = max_length

        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        for cell in worksheet[2]:
            cell.fill = yellow_fill

    return template_buffer.getvalue()


def records_from_editor(standard_presets):
    st.info("点击左上角第一个单元格，然后按 Ctrl+V，可直接粘贴 Excel 数据。")
    initial_df = pd.DataFrame(
        {
            "序号": range(1, 101),
            "证书编号": [""] * 100,
            "姓名": [""] * 100,
            "身份证号": [""] * 100,
            "培训日期": [""] * 100,
            "标准号": [""] * 100,
        }
    )
    edited_df = st.data_editor(
        initial_df,
        num_rows="fixed",
        width="stretch",
        hide_index=True,
        height=380,
        column_config={
            "序号": st.column_config.NumberColumn("序号", width=40, disabled=True),
            "证书编号": st.column_config.TextColumn("证书编号", width="small"),
            "姓名": st.column_config.TextColumn("姓名", width="small"),
            "身份证号": st.column_config.TextColumn("身份证号", width="medium"),
            "培训日期": st.column_config.TextColumn("培训日期", width="medium"),
            "标准号": st.column_config.TextColumn(
                "标准号",
                width="large",
                help="可填写：" + "、".join(standard_presets.values()),
            ),
        },
    )

    records = []
    for row in edited_df.drop(columns=["序号"]).to_dict("records"):
        cleaned = {key: clean_value(value) for key, value in row.items()}
        if any(cleaned.values()):
            records.append(cleaned)
    return records


def records_from_upload():
    left_column, right_column = st.columns([2, 3])
    with left_column:
        st.download_button(
            label="📥 下载标准模板",
            data=make_excel_template(),
            file_name="学员信息上传模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
        st.caption("系统会自动跳过标有“示例”的行。")

    with right_column:
        uploaded_data = st.file_uploader(
            "上传学员信息文件",
            type=["xlsx", "csv"],
            label_visibility="collapsed",
        )

    if not uploaded_data:
        return []

    try:
        if uploaded_data.name.lower().endswith(".csv"):
            data_frame = pd.read_csv(uploaded_data, dtype=str).fillna("")
        else:
            data_frame = pd.read_excel(uploaded_data, dtype=str).fillna("")
    except Exception as error:
        st.error(f"无法读取文件：{error}")
        return []

    if "姓名" not in data_frame.columns:
        st.error("文件中缺少“姓名”列，请使用标准模板。")
        return []

    records = []
    for row in data_frame.to_dict("records"):
        cleaned = {key: clean_value(value) for key, value in row.items()}
        if "示例" in cleaned.get("姓名", "") or "示例" in cleaned.get("证书编号", ""):
            continue
        if cleaned.get("姓名"):
            records.append(cleaned)

    if records:
        st.success(f"已成功加载 {len(records)} 条有效数据。")
    else:
        st.warning("文件中没有可生成的有效姓名。")
    return records


def quick_records(standard_presets):
    selected_labels = []
    standard_labels = list(standard_presets.keys())
    with st.container(border=True, key="quick_setup"):
        section_title("① 选择标准与培训日期")
        selected_count = sum(
            1
            for label in standard_labels
            if st.session_state.get(f"standard_option_{label}")
        )
        popover_label = (
            f"已选择 {selected_count} 个标准"
            if selected_count
            else "选择标准（支持多选）"
        )
        standard_column, date_column = st.columns(2, gap="medium")

        with standard_column:
            st.markdown('<p class="field-label">培训标准</p>', unsafe_allow_html=True)
            with st.popover(popover_label, width="stretch"):
                st.caption("勾选完成后，点击弹出框外即可关闭。")
                for row_start in range(0, len(standard_labels), 2):
                    standard_columns = st.columns(2)
                    for column_index, label in enumerate(
                        standard_labels[row_start : row_start + 2]
                    ):
                        with standard_columns[column_index]:
                            if st.checkbox(label, key=f"standard_option_{label}"):
                                selected_labels.append(label)

        with date_column:
            st.markdown(
                '<p class="field-label">培训日期 '
                '<span class="field-hint">单日点一次，区间再点结束日</span></p>',
                unsafe_allow_html=True,
            )
            selected_dates = st.date_input(
                "培训日期",
                value=[],
                format="YYYY/MM/DD",
                label_visibility="collapsed",
                key="quick_training_dates",
            )

        custom_standard = st.text_input(
            "自定义标准（可选）",
            placeholder="例如：企业内部标准 Q/ABC 001-2026",
            key="quick_custom_standard",
        )

        standard_values = [standard_presets[label] for label in selected_labels]
        if custom_standard:
            standard_values.extend(re.split(r"[\r\n,，;；]+", custom_standard))
        standard_values = merge_unique(standard_values)
        standards = "、".join(standard_values)

        if len(selected_dates) == 1:
            training_date = format_training_date(selected_dates[0])
        elif len(selected_dates) == 2:
            training_date = format_training_date(
                selected_dates[0], selected_dates[1]
            )
        else:
            training_date = ""

        standard_status = (
            standard_chips_html(standard_values)
            if standard_values
            else '<div class="setup-v is-pending">尚未选择</div>'
        )
        date_status = (
            f'<div class="setup-v">{html.escape(training_date)}</div>'
            if training_date
            else '<div class="setup-v is-pending">尚未选择</div>'
        )
        st.markdown(
            f"""
            <div class="setup-status">
                <div>
                    <div class="setup-k">写入证书的标准</div>
                    {standard_status}
                </div>
                <div>
                    <div class="setup-k">写入证书的日期</div>
                    {date_status}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if standard_values:
            st.button(
                "清空已选标准",
                key="clear_quick_standards",
                type="tertiary",
                on_click=clear_quick_standards,
                args=(standard_labels,),
            )

    with st.container(border=True, key="quick_people"):
        section_title("② 粘贴学员信息")
        st.markdown(
            '<p class="field-label">学员信息 '
            '<span class="field-hint">每行一个姓名，或从 Excel 直接复制“姓名、身份证号”两列</span></p>',
            unsafe_allow_html=True,
        )
        raw_people = st.text_area(
            "学员信息",
            height=180,
            placeholder=(
                "张三\n李四\n王五\n\n"
                "也可以从 Excel 复制两列后直接粘贴"
            ),
            label_visibility="collapsed",
            key="quick_people_input",
        )
        people = parse_people(raw_people)

        id_count = sum(bool(person["身份证号"]) for person in people)
        if people:
            count_column, clear_people_column = st.columns(
                [4, 1], vertical_alignment="center"
            )
            with count_column:
                st.markdown(
                    '<div class="status-line">'
                    f"已识别 {len(people)} 人，其中 {id_count} 人包含身份证号"
                    "</div>",
                    unsafe_allow_html=True,
                )
            with clear_people_column:
                st.button(
                    "清空人员",
                    key="clear_quick_people",
                    type="tertiary",
                    on_click=clear_people_input,
                    width="stretch",
                )

            preview_people = [
                {
                    "姓名": person["姓名"],
                    "身份证号": mask_id_card(person["身份证号"]),
                }
                for person in people[:10]
            ]
            st.dataframe(
                preview_people,
                width="stretch",
                hide_index=True,
                height=min(390, 38 + 35 * len(preview_people)),
            )
            if len(people) > len(preview_people):
                st.caption(f"当前仅预览前 10 人，其余 {len(people) - 10} 人已识别。")

        st.caption("身份证号仅用于本次生成，中间 8 位自动隐藏，应用不保存任何数据。")

        with st.expander("自动生成证书编号（可选）"):
            prefix_column, start_column, digits_column = st.columns([2, 1, 1])
            with prefix_column:
                number_prefix = st.text_input(
                    "编号前缀",
                    placeholder="例如：T-2026-",
                    help="不填写则证书编号留空。",
                )
            with start_column:
                number_start = st.number_input(
                    "起始序号", min_value=0, value=1, step=1
                )
            with digits_column:
                number_digits = st.number_input(
                    "序号位数", min_value=1, max_value=8, value=3, step=1
                )
            if number_prefix.strip():
                sample = f"{number_prefix.strip()}{int(number_start):0{int(number_digits)}d}"
                st.caption(f"第一位学员的编号将是：{sample}")

    records = []
    for index, person in enumerate(people):
        certificate_number = ""
        if number_prefix.strip():
            sequence = int(number_start) + index
            certificate_number = f"{number_prefix.strip()}{sequence:0{int(number_digits)}d}"
        records.append(
            {
                "证书编号": certificate_number,
                "姓名": person["姓名"],
                "身份证号": person["身份证号"],
                "培训日期": training_date,
                "标准号": standards,
            }
        )

    return records, {
        "has_standard": bool(standards),
        "standard_count": len(standard_values),
        "has_date": bool(training_date),
        "training_date": training_date,
        "people_count": len(people),
        "id_count": id_count,
    }


st.set_page_config(page_title="证书智能制作工具", page_icon="🎓", layout="centered")
st.markdown(
    """
    <style>
    :root {
        --ink: #17382f;
        --ink-soft: #3d5c54;
        --muted: #6f8580;
        --brand: #176b57;
        --brand-deep: #125647;
        --brand-tint: #e6f3ee;
        --line: #d9e6e1;
        --line-strong: #c9dbd4;
        --surface: #ffffff;
        --surface-soft: #f3f8f6;
    }
    header[data-testid="stHeader"] {
        background: transparent;
    }
    .stApp {
        background:
            radial-gradient(1100px 380px at 50% -120px, #dcefe7 0%, rgba(248, 250, 249, 0) 72%),
            #f8faf9;
    }
    .block-container {
        max-width: 960px;
        padding-top: 3.4rem;
        padding-bottom: 4.5rem;
    }
    .hero {
        padding: 0.4rem 0.2rem 1.5rem;
    }
    .hero-eyebrow {
        color: var(--brand);
        font-size: 0.74rem;
        font-weight: 700;
        letter-spacing: 0.14em;
        margin-bottom: 0.45rem;
    }
    .hero h1 {
        color: var(--ink);
        font-size: 2.05rem;
        line-height: 1.2;
        letter-spacing: -0.02em;
        margin: 0 0 0.5rem;
        padding: 0;
    }
    .hero p {
        color: var(--muted);
        font-size: 1rem;
        margin: 0;
    }
    /* Bordered containers that start with a panel title become cards. */
    [data-testid="stVerticalBlock"]:has(> [data-testid="stElementContainer"] .panel-title) {
        gap: 0.7rem;
        padding: 1.35rem 1.5rem 1.45rem !important;
        border: 1px solid var(--line) !important;
        border-radius: 16px !important;
        background: var(--surface);
        box-shadow: 0 1px 2px rgba(23, 56, 47, 0.04), 0 10px 28px rgba(23, 56, 47, 0.05);
    }
    .panel-title {
        color: var(--ink);
        font-size: 1.12rem;
        font-weight: 700;
        letter-spacing: -0.015em;
        margin: 0 0 0.35rem;
    }
    [data-testid="stMarkdownContainer"]:has(.panel-title),
    [data-testid="stMarkdownContainer"]:has(.field-label) {
        padding-bottom: 0;
    }
    .field-label {
        color: var(--ink-soft);
        font-size: 0.8rem;
        font-weight: 650;
        letter-spacing: 0.03em;
        margin: 0 0 0.3rem;
    }
    .field-hint {
        margin-left: 0.45rem;
        color: var(--muted);
        font-size: 0.74rem;
        font-weight: 500;
        letter-spacing: 0;
    }
    /* Unify every text-like control: white, hairline border, 44px tall. */
    [data-testid="stTextInputRootElement"],
    [data-testid="stNumberInputContainer"],
    [data-testid="stDateInputField"],
    [data-testid="stTextAreaRootElement"],
    div[data-testid="stPopover"] > button {
        background-color: var(--surface) !important;
        border: 1px solid var(--line-strong) !important;
        border-radius: 10px !important;
        box-shadow: none !important;
        transition: border-color 160ms ease, box-shadow 160ms ease;
    }
    [data-testid="stTextInputRootElement"],
    [data-testid="stNumberInputContainer"],
    [data-testid="stDateInputField"],
    div[data-testid="stPopover"] > button {
        min-height: 2.75rem !important;
    }
    [data-testid="stTextInputRootElement"]:hover,
    [data-testid="stNumberInputContainer"]:hover,
    [data-testid="stDateInputField"]:hover,
    [data-testid="stTextAreaRootElement"]:hover,
    div[data-testid="stPopover"] > button:hover {
        border-color: var(--brand) !important;
    }
    [data-testid="stTextInputRootElement"]:focus-within,
    [data-testid="stNumberInputContainer"]:focus-within,
    [data-testid="stDateInputField"]:focus-within,
    [data-testid="stTextAreaRootElement"]:focus-within {
        border-color: var(--brand) !important;
        box-shadow: 0 0 0 3px rgba(23, 107, 87, 0.14) !important;
    }
    div[data-testid="stPopover"] > button {
        justify-content: space-between;
        color: var(--ink) !important;
        font-weight: 500;
    }
    div[data-testid="stPopover"] > button:hover {
        background: #f7fbf9 !important;
    }
    [data-testid="stTextAreaRootElement"] textarea {
        line-height: 1.6;
    }
    .setup-status {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 1rem;
        margin: 0.2rem 0 0;
        padding: 0.85rem 1rem;
        border-radius: 12px;
        background: var(--surface-soft);
    }
    .setup-k {
        color: var(--muted);
        font-size: 0.7rem;
        font-weight: 700;
        letter-spacing: 0.08em;
        margin-bottom: 0.3rem;
    }
    .setup-v {
        color: var(--ink);
        font-size: 0.95rem;
        font-weight: 650;
        line-height: 1.4;
    }
    .setup-v.is-pending {
        color: var(--muted);
        font-weight: 500;
    }
    .chip-row {
        display: flex;
        flex-wrap: wrap;
        gap: 0.4rem;
    }
    .standard-chip {
        display: inline-flex;
        align-items: center;
        padding: 0.28rem 0.6rem;
        border-radius: 999px;
        color: var(--brand-deep);
        background: var(--brand-tint);
        font-size: 0.82rem;
        font-weight: 600;
        line-height: 1.25;
    }
    .status-line {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        color: var(--brand-deep);
        font-size: 0.9rem;
        font-weight: 600;
    }
    .status-line::before {
        content: "";
        width: 8px;
        height: 8px;
        border-radius: 50%;
        background: #1f8a6e;
        flex: none;
    }
    .check-card {
        display: grid;
        grid-template-columns: repeat(2, minmax(0, 1fr));
        gap: 0.6rem 1rem;
        padding: 0.9rem 1rem;
        margin: 0.15rem 0 0.85rem;
        border-radius: 12px;
        background: var(--surface-soft);
    }
    .check-row {
        display: flex;
        gap: 0.5rem;
        align-items: flex-start;
        font-size: 0.9rem;
        line-height: 1.45;
    }
    .check-row > span:first-child {
        width: 1.1rem;
        flex: none;
        text-align: center;
        font-weight: 700;
    }
    .check-ok { color: var(--brand-deep); }
    .check-pending { color: #8a6741; }
    div[data-testid="stButton"] button[kind="primary"] {
        min-height: 3.1rem;
        font-size: 1rem;
        font-weight: 700;
        border-radius: 12px;
        box-shadow: 0 6px 18px rgba(23, 107, 87, 0.18);
    }
    div[data-testid="stButton"] button[kind="primary"]:disabled {
        background: var(--surface-soft);
        border-color: var(--line);
        color: var(--muted);
        box-shadow: none;
    }
    div[data-testid="stButton"] button[kind="tertiary"] {
        color: var(--muted);
        padding-left: 0.2rem;
        padding-right: 0.2rem;
    }
    div[data-testid="stButton"] button[kind="tertiary"]:hover {
        color: #a33e3e;
        background: transparent;
    }
    [data-testid="stExpander"] details {
        border-color: var(--line) !important;
        border-radius: 12px !important;
    }
    [data-testid="stExpander"] summary {
        color: var(--ink-soft);
        font-weight: 600;
    }
    [data-testid="stFileUploaderDropzone"] {
        border: 1px dashed var(--line-strong) !important;
        background: var(--surface-soft) !important;
        border-radius: 12px !important;
    }
    div[data-testid="stDownloadButton"] button {
        min-height: 2.9rem;
        border-radius: 12px;
        font-weight: 600;
    }
    @media (max-width: 640px) {
        .block-container { padding-top: 2.4rem; }
        .hero h1 { font-size: 1.6rem; }
        .check-card,
        .setup-status { grid-template-columns: 1fr; }
    }
    </style>
    <div class="hero">
        <div class="hero-eyebrow">CERTIFICATE GENERATOR</div>
        <h1>内审员证书智能制作工具</h1>
        <p>选择标准和培训日期，粘贴学员信息，一次生成全部证书。</p>
    </div>
    """,
    unsafe_allow_html=True,
)

standard_presets = load_standard_presets()
data_to_process = []
inputs_valid = False
quick_summary = None

mode = (
    st.segmented_control(
        "录入方式",
        ["快速生成", "完整信息录入"],
        default="快速生成",
        help="大多数场景使用快速生成；每名学员信息不同时使用完整信息录入。",
    )
    or "快速生成"
)

if mode == "快速生成":
    data_to_process, quick_summary = quick_records(standard_presets)
    inputs_valid = (
        quick_summary["has_standard"]
        and quick_summary["has_date"]
        and bool(data_to_process)
    )
else:
    with st.container(border=True, key="full_entry"):
        section_title("① 录入完整证书信息")
        full_mode = (
            st.segmented_control(
                "完整信息录入方式",
                ["Excel 文件上传", "网页表格填写（支持粘贴）"],
                default="Excel 文件上传",
                label_visibility="collapsed",
            )
            or "Excel 文件上传"
        )
        if full_mode == "Excel 文件上传":
            data_to_process = records_from_upload()
        else:
            data_to_process = records_from_editor(standard_presets)
    inputs_valid = bool(data_to_process)

with st.container(border=True, key="generate_panel"):
    final_step = "③" if mode == "快速生成" else "②"
    section_title(f"{final_step} 确认并生成")

    if os.path.exists(DEFAULT_TEMPLATE):
        template_source = DEFAULT_TEMPLATE
        template_label = "内置模板：内审员证书.docx"
        uploaded_template = st.session_state.get("custom_word_template")
        if uploaded_template:
            template_source = uploaded_template
            template_label = f"自定义模板：{uploaded_template.name}"
    else:
        template_source = st.file_uploader(
            "上传 Word 模板", type=["docx"], key="required_word_template"
        )
        template_label = (
            f"自定义模板：{template_source.name}"
            if template_source
            else "尚未选择 Word 模板"
        )
        st.warning("仓库中未发现默认模板，请上传 Word 模板。")

    template_ready = bool(template_source)
    if mode == "快速生成":
        checklist = [
            (
                quick_summary["has_standard"],
                f"已选择 {quick_summary['standard_count']} 个标准"
                if quick_summary["has_standard"]
                else "尚未选择标准",
            ),
            (
                quick_summary["has_date"],
                f"培训日期：{quick_summary['training_date']}"
                if quick_summary["has_date"]
                else "尚未选择培训日期",
            ),
            (
                bool(data_to_process),
                (
                    f"已录入 {quick_summary['people_count']} 人，"
                    f"含身份证号 {quick_summary['id_count']} 人"
                )
                if data_to_process
                else "尚未粘贴学员信息",
            ),
            (template_ready, template_label),
        ]
    else:
        checklist = [
            (
                bool(data_to_process),
                f"已载入 {len(data_to_process)} 条完整信息"
                if data_to_process
                else "尚未载入学员信息",
            ),
            (template_ready, template_label),
        ]
    render_checklist(checklist)

    if os.path.exists(DEFAULT_TEMPLATE):
        with st.expander("更换为其他 Word 模板"):
            st.file_uploader(
                "上传自定义 Word 模板",
                type=["docx"],
                key="custom_word_template",
                help="模板中可使用 {{ number }}、{{ name }}、{{ id_card }}、{{ date }}、{{ standards }} 变量。",
            )

    can_generate = template_ready and inputs_valid
    button_label = (
        f"🚀 开始生成 {len(data_to_process)} 份证书"
        if data_to_process
        else "🚀 开始批量生成"
    )
    if st.button(
        button_label,
        type="primary",
        width="stretch",
        disabled=not can_generate,
    ):
        try:
            progress_bar = st.progress(0, text="正在生成证书……")
            template_bytes = read_template_bytes(template_source)
            result = generate_documents(
                template_bytes,
                data_to_process,
                progress_callback=lambda value: progress_bar.progress(
                    value, text="正在生成证书……"
                ),
            )
            progress_bar.empty()
            if not result:
                st.error("没有找到可生成证书的有效姓名。")
            else:
                st.session_state["generation_result"] = result
                st.toast(f"已生成 {result['count']} 份证书", icon="✅")
        except Exception as error:
            st.error(f"制作失败：{error}")

    if not can_generate:
        pending_items = [text for complete, text in checklist if not complete]
        st.caption("还需完成：" + "；".join(pending_items))

result = st.session_state.get("generation_result")
if result:
    with st.container(border=True, key="result_panel"):
        section_title("证书生成完成")
        st.markdown(
            f'<div class="status-line">共生成 {result["count"]} 份证书，选择需要的下载方式</div>',
            unsafe_allow_html=True,
        )
        download_word, download_pdf, download_zip = st.columns(3)
        with download_word:
            st.download_button(
                "下载合并 Word",
                data=result["merged"],
                file_name="证书汇总导出.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                width="stretch",
            )
        with download_pdf:
            if result.get("pdf"):
                st.download_button(
                    "下载合并 PDF",
                    data=result["pdf"],
                    file_name="证书汇总导出.pdf",
                    mime="application/pdf",
                    width="stretch",
                )
            else:
                st.button("下载合并 PDF", disabled=True, width="stretch")
                st.caption("本次未能生成 PDF，请改用 Word。")
        with download_zip:
            st.download_button(
                "下载单独证书 ZIP",
                data=result["zip"],
                file_name="单独证书.zip",
                mime="application/zip",
                width="stretch",
                help="ZIP 内同时包含每人的 Word 和 PDF。",
            )

